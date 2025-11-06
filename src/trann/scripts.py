import os

ERROR = '\033[91m[ERROR]\033[0m'

class TableReader:
    """
    A universal class for reading tabular data from CSV or Excel files.

    The class automatically determines the file extension and selects
    the appropriate reader library. For CSV files, it can also detect
    the delimiter automatically (comma, semicolon, tab, space, etc.).

    Usage example:
        >>> reader = TableReader("data.xlsx")
        >>> rows = reader.read()
        >>> for row in rows:
        ...     print(row)
    """
    def __init__(self, filepath):
        """
        Initialize a TableReader instance.

        :param filepath: str, Path to the table file (.csv, .xls, or .xlsx)
        """
        self.filepath = filepath
        self.ext = os.path.splitext(filepath)[1].lower()

    def read(self):
        """
        Read the table based on the detected file extension.

        This method automatically chooses the appropriate reading
        function based on the file extension.

        :return: list[list[Any]]: A list of rows representing the table data
        :raises ValueError: If the file extension is not supported
        """
        if self.ext == ".csv":
            return self._read_csv()
        elif self.ext in (".xls", ".xlsx"):
            return self._read_xlsx()
        else:
            raise ValueError(f"{ERROR} Unsupported extension: {self.ext}")

    def _read_csv(self):
        """
        Read a CSV file with automatic delimiter detection.

        This method uses ``csv.Sniffer`` to detect delimiters among
        common characters (comma, semicolon, space, tab). If detection
        fails, it falls back to the default Excel dialect (comma).

        :return: list[list[str]]: Table rows as a list of string lists
        :notes: This method reads the entire file into memory, so it may
                not be suitable for extremely large CSVs.
        Returns
        -------

        """
        import csv
        with open(self.filepath, newline="", encoding="utf-8") as f:
            sample = f.read(2048)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";, \t")
            except csv.Error:
                dialect = csv.get_dialect("excel")  # стандартный: запятая
            reader = csv.reader(f, dialect)
            return list(reader)

    def _read_xlsx(self):
        """
        Read an Excel (.xlsx or .xls) file using the openpyxl library.

        This method reads the active sheet and returns its content
        as a list of lists, where each inner list represents one row.

        :return: list[list[Any]]: Sheet data as a list of rows
        """

        from openpyxl import load_workbook
        wb = load_workbook(self.filepath)
        ws = wb.active
        return [list(row.split()) for row in ws.iter_rows(values_only=True)]


def get_paths(args):
    """
    Resolve and validate paths to the information (.csv) and tree (.tree) files.

    This function checks whether the user provided explicit paths via CLI
    arguments (`args.info`, `args.tree`). If not, it automatically searches
    the current working directory for files matching the expected extensions
    (``.csv`` for the info file and ``.tree`` for the tree file).

    If a specified or discovered file does not exist, a formatted error
    message is printed and the function returns 1.

    :param args: argparse.Namespace, An object containing at least two
                 attributes: ``info`` (path to a .csv file or '.') and
                 ``tree`` (path to a .tree file or '.')
    :return: tuple[str, str] | int:
             - On success: a tuple ``(info_path, tree_path)`` with absolute paths.
             - On failure: returns ``1`` if any required file is missing
               or not found automatically.
    :notes:
        - If ``args.info`` equals '.', the function searches for the first
          ``.csv`` file in the current working directory.
        - If ``args.tree`` equals '.', the function searches for the first
          ``.tree`` file in the current working directory.
        - Printed error messages use the global ``ERROR`` variable for styling.
    """
    if args.info != '.':
        info_path = args.info
        if not os.path.exists(info_path):
            print(f'{ERROR} File {info_path} is not exists')
            return 1
    else:
        try:
            info_path = f'{os.getcwd()}/{[i for i in os.listdir(os.getcwd()) if i.endswith(".csv")][0]}'

        except IndexError:
            print(f'{ERROR} There are no files similar to the table in {os.getcwd()}')
            return 1

    if args.tree != '.':
        tree_path = args.tree
        if not os.path.exists(tree_path):
            print(f'{ERROR} File {tree_path} is not exists')
            return 1
    else:
        try:
            tree_path = f'{os.getcwd()}/{[i for i in os.listdir(os.getcwd()) if i.endswith(".tree")][0]}'
        except IndexError:
            print(f'{ERROR} There are no files similar to the tree in {os.getcwd()}')
            return 1

    return info_path, tree_path
